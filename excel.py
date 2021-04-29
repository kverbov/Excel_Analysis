from openpyxl import Workbook, load_workbook
import os

class Excel:
    workbook=0

    def __init__(self):
        self.workbook=0

    def writeBook(self, fpath):
        pass

    def

    def sortFile(self, fpath, resultPath):

        try:
            f = open(fpath,'r')
        except IOError as e:
            print(e)
            return False

        print('Открыли  файл '+ fpath)
        mas=0
        cnt=0
        hint=102400
        while True:
            try:
                mas=f.readlines(hint)
                if(len(mas)==0):
                    print('Конец файла')
                    break
            except Exception as e:
                print('Чтение файла - достигнут конец '+fpath+'\n'+e)
                break

            frez = open(resultPath, 'a')
            # Разбиваем строки на массив
            rez=[]
            for item in mas:
                if(item.find('п»ї')>-1):
                    continue
                cnt=cnt+1
                line = item.replace('\n', '').rsplit(';')
                line[2]=line[2].replace(',','.')
                if(not line[2].isalpha() and float(line[2])>97):
                    rez.append(item)
                else:
                    continue

            # запичсываем в файл построчно
            frez.write('\n')
            for line in rez:
                frez.writelines(line)
            frez.close()
            #wb = load_workbook(fpath)
            #wb[2]
        f.close()
        print('Обработано '+str(cnt) +' строк')

        def saveBook(self, workbook, fpath):
            wb = Workbook()
            # grab the active worksheet
            ws = wb.active

            # Data can be assigned directly to cells
            # ws['A1'] = 42

            # Rows can also be appended
            # ws.append([1, 2, 3])

            # Save the file
            return True

    def saveBook(self,fpath):
        self.workbook.save(fpath)
